#!/usr/bin/env python3
"""
Run SBOM generation for all CNCF projects listed in an Excel file.

This script reads the GitHub repository URLs from the spreadsheet
and calls sbom_generate.py for each project.

Author: Gayathri
"""

import subprocess
from pathlib import Path

from openpyxl import load_workbook


########################################################################
# Configuration
########################################################################

# Excel file containing CNCF projects.
BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILE = BASE_DIR / "Copy of cncf_projects.xlsx"

# Name of your SBOM generator script.
SBOM_SCRIPT = "sbom_generate.py"


########################################################################
# Read Excel File
########################################################################

workbook = load_workbook(EXCEL_FILE)

worksheet = workbook["CNCF Projects"]

########################################################################
# Find Required Columns
########################################################################

headers = {}

for column in worksheet[1]:

    headers[column.value] = column.column

project_column = headers["Project Name"]

github_column = headers["GitHub Repository URL"]

sbomit_column = headers["SBOMit"]

########################################################################
# Collect Projects
########################################################################

projects = []

for row in worksheet.iter_rows(min_row=2):

    project_name = row[project_column - 1].value

    github_url = row[github_column - 1].value

    sbomit_status = row[sbomit_column - 1].value

    if sbomit_status == "T" and github_url:

        projects.append(

            (project_name, github_url)

        )

########################################################################
# Run SBOM Generator
########################################################################

total_projects = len(projects)

print(f"\nFound {total_projects} projects.\n")

for index, (project_name, github_url) in enumerate(projects, start=1):

    print("=" * 70)

    print(

        f"[{index}/{total_projects}] "

        f"{project_name}"

    )

    try:

        subprocess.run(

            [

                "python",

                SBOM_SCRIPT,

                github_url,

                "--syft",

                "--trivy",

                "--spdx"

            ],

            check=True

        )

        print("Completed successfully.\n")

    except subprocess.CalledProcessError:

        print(

            "SBOM generation failed.\n"

        )

print("=" * 70)

print("Finished processing all projects.")